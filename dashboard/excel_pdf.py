import io
import datetime
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def download_excel_dynamic(data, columns, filename_prefix="export"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD") 
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write header row with styles
    headers = [header for field, header in columns]
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(data, start=2): 
        for col_idx, (field, header) in enumerate(columns, start=1):
            value = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

    # --- Auto adjust column width dynamically ---
    for col_idx, (field, header) in enumerate(columns, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Add some padding
        adjusted_width = max_length + 3
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare file in memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



def generate_pdf_dynamic(data, columns, filename_prefix="export", html=None):
    """
    Generates a PDF file in memory from provided data and columns (using an HTML template), returns HttpResponse for download.
    :param data: list[dict]
    :param columns: list of (field, header)
    :param filename_prefix: prefix for the generated filename
    :param html: Optional rendered HTML string. If provided, use this as the html to render PDF.
    """
    from django.template.loader import render_to_string
    from weasyprint import HTML

    if html is None:
        headers = [header for field, header in columns]
        rows = []
        for row in data:
            row_data = [row.get(field, "") for field, header in columns]
            rows.append(row_data)

        context = {
            'headers': headers,
            'rows': rows,
            'title': filename_prefix.capitalize()
        }

        html_string = render_to_string('pdf/pdf_table_template.html', context)
    else:
        html_string = html

    # Generate the PDF
    pdf_file = io.BytesIO()
    HTML(string=html_string).write_pdf(target=pdf_file)
    pdf_file.seek(0)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.pdf"
    response = HttpResponse(
        pdf_file.read(),
        content_type="application/pdf"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

